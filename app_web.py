#!/usr/bin/env python3
"""
Generador de Horarios Semanales / Anuales — Versión Web
========================================================
Columna "Dia" en Actividades:
  • Toda la semana → aparece en el horario principal (Lun–Dom)
  • Domingo        → horario independiente solo para domingos
"""

import streamlit as st
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import random
from datetime import datetime, timedelta
from collections import defaultdict
from io import BytesIO
import calendar as cal_mod


# ─────────────────────────────────────────────────────────────────────────────
# Constantes
# ─────────────────────────────────────────────────────────────────────────────

DAYS = ['Lunes', 'Martes', 'Miércoles', 'Jueves', 'Viernes', 'Sábado', 'Domingo']

MONTHS_ES = [
    'Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio',
    'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre'
]

RESPONSABLE_NAMES = frozenset({'responsable 1', 'responsable 2'})


# ─────────────────────────────────────────────────────────────────────────────
# Utilidades de fecha  (semana = lunes → domingo, 7 días)
# ─────────────────────────────────────────────────────────────────────────────

def get_week_dates(reference_date=None):
    """7 tuplas (día, 'dd/mm/yyyy') lunes–domingo de la semana de reference_date."""
    if reference_date is None:
        reference_date = datetime.now()
    monday = reference_date - timedelta(days=reference_date.weekday())
    return [(DAYS[i], (monday + timedelta(days=i)).strftime('%d/%m/%Y'))
            for i in range(7)]


def get_year_weeks(year=None):
    """Lista de (monday_date, week_dates_7) para cada semana del año."""
    if year is None:
        year = datetime.now().year
    jan1 = datetime(year, 1, 1)
    first_monday = jan1 - timedelta(days=jan1.weekday())
    if first_monday.year < year:
        first_monday += timedelta(weeks=1)
    weeks, current = [], first_monday
    while current.year == year:
        week_dates = [(DAYS[i], (current + timedelta(days=i)).strftime('%d/%m/%Y'))
                      for i in range(7)]
        weeks.append((current, week_dates))
        current += timedelta(weeks=1)
    return weeks


def get_month_weeks(year: int, month: int):
    """
    Semanas que se solapan con month/year.
    Retorna (monday_date, week_dates_7, week_offset_global).
    """
    year_weeks   = get_year_weeks(year)
    week_mondays = {md: i for i, (md, _) in enumerate(year_weeks)}

    first_day    = datetime(year, month, 1)
    first_monday = first_day - timedelta(days=first_day.weekday())
    last_day_num = cal_mod.monthrange(year, month)[1]
    last_day     = datetime(year, month, last_day_num)

    result, current = [], first_monday
    while current <= last_day:
        week_dates = [(DAYS[i], (current + timedelta(days=i)).strftime('%d/%m/%Y'))
                      for i in range(7)]
        offset = week_mondays.get(current, 0)
        result.append((current, week_dates, offset))
        current += timedelta(weeks=1)
    return result


# ─────────────────────────────────────────────────────────────────────────────
# Normalización
# ─────────────────────────────────────────────────────────────────────────────

def normalize_sex(value):
    if pd.isna(value): return 'Indiferente'
    v = str(value).strip().lower()
    if v in ('h', 'hombre', 'm', 'masculino', 'male', 'masc', 'hom'): return 'Hombre'
    if v in ('mujer', 'f', 'femenino', 'female', 'fem', 'w', 'muj'):  return 'Mujer'
    return 'Indiferente'


def normalize_dia(value):
    """'Toda la semana' | 'Domingo'"""
    if pd.isna(value): return 'Toda la semana'
    v = str(value).strip().lower()
    if 'domingo' in v or v in ('dom', 'sun', 'sunday'): return 'Domingo'
    return 'Toda la semana'


def student_can_do(student_sex: str, activity_sex: str) -> bool:
    return activity_sex == 'Indiferente' or student_sex == activity_sex


def opposite_sex(sex: str) -> str:
    return {'Hombre': 'Mujer', 'Mujer': 'Hombre'}.get(sex, sex)


def effective_sex(act_name: str, original_sex: str,
                  week_offset: int, day_idx: int) -> str:
    """Alterna Responsable 1/2: paridad = (week_offset + day_idx) % 2."""
    if act_name.strip().lower() not in RESPONSABLE_NAMES:
        return original_sex
    if original_sex == 'Indiferente':
        return original_sex
    return original_sex if (week_offset + day_idx) % 2 == 0 else opposite_sex(original_sex)


# ─────────────────────────────────────────────────────────────────────────────
# Preparación de datos
# ─────────────────────────────────────────────────────────────────────────────

def _prepare_dfs(activities_df, students_df):
    """
    Normaliza y separa actividades por columna 'Dia'.
    Retorna (acts_semana, acts_domingo, students) — listas de dicts.
    Si no existe la columna 'Dia', todo va a 'Toda la semana'.
    """
    acts = activities_df.copy()
    stus = students_df.copy()
    acts['Sexo'] = acts['Sexo'].apply(normalize_sex)
    stus['Sexo'] = stus['Sexo'].apply(normalize_sex)

    if 'Dia' in acts.columns:
        acts['Dia'] = acts['Dia'].apply(normalize_dia)
    else:
        acts['Dia'] = 'Toda la semana'

    acts_semana  = acts[acts['Dia'] == 'Toda la semana'].to_dict('records')
    acts_domingo = acts[acts['Dia'] == 'Domingo'].to_dict('records')
    return acts_semana, acts_domingo, stus.to_dict('records')


# ─────────────────────────────────────────────────────────────────────────────
# Algoritmo de generación
# ─────────────────────────────────────────────────────────────────────────────

def _assign_day(activities, available, history, week_offset, day_idx, schedule_day):
    """Asigna actividades a alumnos para un día. Modifica schedule_day y history."""
    def eff(a):
        return effective_sex(a['Nombre actividad'], a['Sexo'], week_offset, day_idx)

    specific = [a for a in activities if eff(a) != 'Indiferente']
    general  = [a for a in activities if eff(a) == 'Indiferente']
    random.shuffle(specific); random.shuffle(general)

    for activity in specific + general:
        act_name = activity['Nombre actividad']
        act_sex  = eff(activity)
        eligible = [sn for sn, ss in available.items() if student_can_do(ss, act_sex)]
        if not eligible:
            schedule_day[act_name] = '—'; continue
        not_done = [s for s in eligible if act_name not in history[s]]
        pool     = not_done if not_done else eligible
        chosen   = random.choice(pool)
        schedule_day[act_name] = chosen
        del available[chosen]
        history[chosen].add(act_name)


def generate_week_schedule(acts_semana: list, students: list,
                           week_offset: int = 0) -> dict:
    """
    schedule[dia][actividad] = alumno (o '—')
    Procesa actividades 'Toda la semana' para los 7 días (Lun–Dom).
    """
    schedule = {day: {} for day in DAYS}
    history  = {s['Nombre Alumnos']: set() for s in students}
    for day_idx, day in enumerate(DAYS):
        available = {s['Nombre Alumnos']: s['Sexo'] for s in students}
        _assign_day(acts_semana, available, history, week_offset, day_idx,
                    schedule[day])
    return schedule


def generate_sunday_schedule(acts_domingo: list, students: list,
                              week_offset: int = 0) -> dict:
    """
    Asigna actividades 'Domingo' para un domingo.
    Retorna {'Domingo': {actividad: alumno}}.
    """
    if not acts_domingo:
        return {'Domingo': {}}
    available = {s['Nombre Alumnos']: s['Sexo'] for s in students}
    history   = {s['Nombre Alumnos']: set() for s in students}
    day_sched = {}
    _assign_day(acts_domingo, available, history, week_offset, 6, day_sched)
    return {'Domingo': day_sched}


# ── Wrappers por periodo ──────────────────────────────────────────────────────

def generate_weekly_all(activities_df, students_df):
    """Retorna (main_sched, sunday_sched | None, acts_semana, acts_domingo)."""
    a_s, a_d, stus = _prepare_dfs(activities_df, students_df)
    main    = generate_week_schedule(a_s, stus, week_offset=0)
    sunday  = generate_sunday_schedule(a_d, stus, week_offset=0) if a_d else None
    return main, sunday, a_s, a_d


def generate_monthly_all(activities_df, students_df, year: int, month: int):
    """Retorna (data_list, acts_semana, acts_domingo)
    data_list: [(monday_date, week_dates, main_sched, sunday_sched|None)]"""
    a_s, a_d, stus = _prepare_dfs(activities_df, students_df)
    result = []
    for mon, wd, offset in get_month_weeks(year, month):
        main   = generate_week_schedule(a_s, stus, week_offset=offset)
        sunday = generate_sunday_schedule(a_d, stus, week_offset=offset) if a_d else None
        result.append((mon, wd, main, sunday))
    return result, a_s, a_d


def generate_annual_all(activities_df, students_df, year=None):
    """Retorna (data_list, acts_semana, acts_domingo)"""
    a_s, a_d, stus = _prepare_dfs(activities_df, students_df)
    result = []
    for offset, (mon, wd) in enumerate(get_year_weeks(year)):
        main   = generate_week_schedule(a_s, stus, week_offset=offset)
        sunday = generate_sunday_schedule(a_d, stus, week_offset=offset) if a_d else None
        result.append((mon, wd, main, sunday))
    return result, a_s, a_d


# ─────────────────────────────────────────────────────────────────────────────
# Estilos Excel
# ─────────────────────────────────────────────────────────────────────────────

C_BLUE_DARK  = "1F4E79"
C_BLUE_MED   = "2E75B6"
C_BLUE_LIGHT = "D6E4F0"
C_WHITE      = "FFFFFF"
C_GRAY_LIGHT = "F5F5F5"
C_SUN_DARK   = "7B3F00"   # marrón/naranja para el Excel de domingos
C_SUN_MED    = "C05621"
C_SUN_LIGHT  = "FDEBD0"

THIN = Border(
    left=Side(style='thin', color='AAAAAA'), right=Side(style='thin', color='AAAAAA'),
    top=Side(style='thin', color='AAAAAA'),  bottom=Side(style='thin', color='AAAAAA'),
)
MED = Border(
    left=Side(style='medium', color='666666'), right=Side(style='medium', color='666666'),
    top=Side(style='medium', color='666666'),  bottom=Side(style='medium', color='666666'),
)

def _fill(h):
    return PatternFill(start_color=h, end_color=h, fill_type='solid')
def _font(bold=False, color="1A1A1A", size=10):
    return Font(bold=bold, color=color, size=size)
def _align(h='center', wrap=True):
    return Alignment(horizontal=h, vertical='center', wrap_text=wrap)


def _write_week_block(ws, start_row: int, week_dates: list,
                      schedule: dict, activity_names: list,
                      show_week_hdr: bool = True,
                      row_h: int = 26, fsz: int = 10,
                      custom_hdr: str = None,
                      hdr_fill: str = C_BLUE_MED,
                      col_fill: str = C_BLUE_DARK) -> int:
    """
    Escribe un bloque semana/domingo en ws desde start_row.
    week_dates: lista de (day_name, date_str) — puede ser 7 días o solo 1 (domingo).
    schedule:   dict[day_name][act_name] = alumno
    Retorna la siguiente fila disponible.
    """
    n_cols   = len(week_dates) + 1
    last_col = get_column_letter(n_cols)
    row      = start_row

    if show_week_hdr:
        hdr_text = custom_hdr or f"Semana del {week_dates[0][1]} al {week_dates[-1][1]}"
        ws.merge_cells(f'A{row}:{last_col}{row}')
        c = ws.cell(row=row, column=1, value=hdr_text)
        c.font = _font(bold=True, color=C_WHITE, size=fsz)
        c.fill = _fill(hdr_fill)
        c.alignment = _align('center')
        ws.row_dimensions[row].height = 20
        row += 1

    # Fila de encabezados de días
    c = ws.cell(row=row, column=1, value="Actividad")
    c.font = _font(bold=True, color=C_WHITE, size=fsz)
    c.fill = _fill(col_fill)
    c.alignment = _align('center')
    c.border = THIN
    ws.row_dimensions[row].height = 40 if fsz >= 10 else 32

    for col_idx, (day, date) in enumerate(week_dates, start=2):
        c = ws.cell(row=row, column=col_idx, value=f"{day}\n{date}")
        c.font = _font(bold=True, color=C_WHITE, size=fsz)
        c.fill = _fill(hdr_fill)
        c.alignment = _align('center')
        c.border = THIN
    row += 1

    # Filas de actividades
    for off, act_name in enumerate(activity_names):
        r        = row + off
        row_fill = C_BLUE_LIGHT if off % 2 == 0 else C_WHITE
        c = ws.cell(row=r, column=1, value=act_name)
        c.font = _font(bold=True, size=fsz)
        c.fill = _fill(row_fill)
        c.alignment = _align('left')
        c.border = THIN
        ws.row_dimensions[r].height = row_h

        for col_idx, (day, _) in enumerate(week_dates, start=2):
            student  = schedule[day].get(act_name, '')
            is_empty = student in ('—', '')
            c = ws.cell(row=r, column=col_idx,
                        value='—' if is_empty else student)
            c.font = _font(size=fsz, color='AAAAAA' if is_empty else '1A1A1A')
            c.fill = _fill(C_GRAY_LIGHT if is_empty else row_fill)
            c.alignment = _align('center')
            c.border = THIN

    return row + len(activity_names) + 1   # +1 fila vacía separadora


def _set_col_widths(ws, n_day_cols: int, act_w: float = 26, day_w: float = 18):
    ws.column_dimensions['A'].width = act_w
    for col in range(2, n_day_cols + 2):
        ws.column_dimensions[get_column_letter(col)].width = day_w


# ─────────────────────────────────────────────────────────────────────────────
# Guardar Excel principal → BytesIO
# ─────────────────────────────────────────────────────────────────────────────

def _build_main_workbook(data_list, act_names_semana, title_str,
                         annual=False, year=None) -> Workbook:
    """
    Construye el workbook del horario principal (Toda la semana, Lun–Dom).
    data_list: [(monday_date, week_dates_7, main_sched, _), ...]
    """
    n_days   = 7
    last_col = get_column_letter(n_days + 1)

    wb = Workbook()

    if annual and year:
        # Una hoja por mes
        months: dict = defaultdict(list)
        for mon, wd, ms, _ in data_list:
            months[mon.month].append((wd, ms))

        first = True
        for month_num in range(1, 13):
            if month_num not in months: continue
            mn = MONTHS_ES[month_num - 1]
            ws = wb.active if first else wb.create_sheet(mn)
            if first: ws.title = mn; first = False
            ws.merge_cells(f'A1:{last_col}1')
            c = ws['A1']
            c.value = f"{mn.upper()} {year}"
            c.font = _font(bold=True, color=C_WHITE, size=13)
            c.fill = _fill(C_BLUE_DARK); c.alignment = _align('center'); c.border = MED
            ws.row_dimensions[1].height = 30
            cur = 2
            for wd, ms in months[month_num]:
                cur = _write_week_block(ws, cur, wd, ms, act_names_semana,
                                        show_week_hdr=True, row_h=22, fsz=9)
            _set_col_widths(ws, n_days, act_w=24, day_w=15)
    else:
        # Hoja única (semanal o mensual)
        ws = wb.active
        ws.title = "Horario"
        ws.merge_cells(f'A1:{last_col}1')
        c = ws['A1']
        c.value = title_str
        c.font = _font(bold=True, color=C_WHITE, size=14)
        c.fill = _fill(C_BLUE_DARK); c.alignment = _align('center'); c.border = MED
        ws.row_dimensions[1].height = 34

        if len(data_list) == 1:
            # Semanal: rango de fechas en fila 2
            _, wd, ms, _ = data_list[0]
            ws.merge_cells(f'A2:{last_col}2')
            c = ws['A2']
            c.value = f"Semana del {wd[0][1]} al {wd[-1][1]}"
            c.font = _font(color=C_WHITE, size=10)
            c.fill = _fill(C_BLUE_MED); c.alignment = _align('center')
            ws.row_dimensions[2].height = 20
            _write_week_block(ws, 3, wd, ms, act_names_semana,
                              show_week_hdr=False, row_h=32, fsz=10)
            _set_col_widths(ws, n_days, act_w=28, day_w=20)
            ws.freeze_panes = 'B4'
        else:
            # Mensual: semanas apiladas
            cur = 2
            for _, wd, ms, _ in data_list:
                cur = _write_week_block(ws, cur, wd, ms, act_names_semana,
                                        show_week_hdr=True, row_h=26, fsz=10)
            _set_col_widths(ws, n_days, act_w=26, day_w=18)

    return wb


def _build_sunday_workbook(data_list, act_names_dom, title_str,
                           annual=False, year=None) -> Workbook:
    """
    Construye el workbook del horario de domingos.
    Cada domingo aparece como una columna de 1 día.
    """
    last_col = get_column_letter(2)   # solo columna A (actividad) + B (domingo)
    wb = Workbook()

    if annual and year:
        months: dict = defaultdict(list)
        for mon, wd, _, ss in data_list:
            if ss: months[mon.month].append((wd, ss))

        first = True
        for month_num in range(1, 13):
            if month_num not in months: continue
            mn = MONTHS_ES[month_num - 1]
            ws = wb.active if first else wb.create_sheet(mn)
            if first: ws.title = mn; first = False
            ws.merge_cells(f'A1:{last_col}1')
            c = ws['A1']
            c.value = f"DOMINGOS — {mn.upper()} {year}"
            c.font = _font(bold=True, color=C_WHITE, size=12)
            c.fill = _fill(C_SUN_DARK); c.alignment = _align('center'); c.border = MED
            ws.row_dimensions[1].height = 28
            cur = 2
            for wd, ss in months[month_num]:
                sun_dates = [wd[6]]   # ('Domingo', 'dd/mm/yyyy')
                cur = _write_week_block(
                    ws, cur, sun_dates, ss, act_names_dom,
                    show_week_hdr=True,
                    custom_hdr=f"Domingo {wd[6][1]}",
                    row_h=22, fsz=9,
                    hdr_fill=C_SUN_MED, col_fill=C_SUN_DARK
                )
            _set_col_widths(ws, 1, act_w=28, day_w=22)
    else:
        ws = wb.active
        ws.title = "Domingos"
        ws.merge_cells(f'A1:{last_col}1')
        c = ws['A1']
        c.value = title_str
        c.font = _font(bold=True, color=C_WHITE, size=14)
        c.fill = _fill(C_SUN_DARK); c.alignment = _align('center'); c.border = MED
        ws.row_dimensions[1].height = 34

        cur = 2
        for _, wd, _, ss in data_list:
            if not ss: continue
            sun_dates = [wd[6]]
            cur = _write_week_block(
                ws, cur, sun_dates, ss, act_names_dom,
                show_week_hdr=True,
                custom_hdr=f"Domingo {wd[6][1]}",
                row_h=30, fsz=10,
                hdr_fill=C_SUN_MED, col_fill=C_SUN_DARK
            )
        _set_col_widths(ws, 1, act_w=30, day_w=24)

    return wb


def _wb_to_bytes(wb: Workbook) -> BytesIO:
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ─────────────────────────────────────────────────────────────────────────────
# API pública de generación + guardado
# ─────────────────────────────────────────────────────────────────────────────

def build_weekly(activities_df, students_df):
    main, sun, a_s, a_d = generate_weekly_all(activities_df, students_df)
    wd = get_week_dates()
    data = [(datetime.now(), wd, main, sun)]

    act_s_names = [a['Nombre actividad'] for a in a_s]
    act_d_names = [a['Nombre actividad'] for a in a_d]

    title = f"HORARIO SEMANAL — {wd[0][1]} al {wd[-1][1]}"
    buf_main = _wb_to_bytes(_build_main_workbook(data, act_s_names, title))
    buf_sun  = (_wb_to_bytes(_build_sunday_workbook(data, act_d_names,
                f"DOMINGOS — {wd[6][1]}"))
                if a_d else None)
    return buf_main, buf_sun, len(a_d) > 0


def build_monthly(activities_df, students_df, year: int, month: int):
    data, a_s, a_d = generate_monthly_all(activities_df, students_df, year, month)
    act_s_names = [a['Nombre actividad'] for a in a_s]
    act_d_names = [a['Nombre actividad'] for a in a_d]
    mn    = MONTHS_ES[month - 1]
    title = f"HORARIO {mn.upper()} {year}"
    buf_main = _wb_to_bytes(_build_main_workbook(data, act_s_names, title))
    buf_sun  = (_wb_to_bytes(_build_sunday_workbook(data, act_d_names,
                f"DOMINGOS — {mn.upper()} {year}"))
                if a_d else None)
    return buf_main, buf_sun, len(a_d) > 0


def build_annual(activities_df, students_df, year: int):
    data, a_s, a_d = generate_annual_all(activities_df, students_df, year)
    act_s_names = [a['Nombre actividad'] for a in a_s]
    act_d_names = [a['Nombre actividad'] for a in a_d]
    buf_main = _wb_to_bytes(_build_main_workbook(
        data, act_s_names, f"HORARIO {year}", annual=True, year=year))
    buf_sun  = (_wb_to_bytes(_build_sunday_workbook(
        data, act_d_names, f"DOMINGOS {year}", annual=True, year=year))
                if a_d else None)
    return buf_main, buf_sun, len(a_d) > 0


# ─────────────────────────────────────────────────────────────────────────────
# Interfaz Web — Streamlit
# ─────────────────────────────────────────────────────────────────────────────

st.set_page_config(
    page_title="Generador de Horarios",
    page_icon="📅",
    layout="centered",
    initial_sidebar_state="collapsed",
)

st.markdown("""
<style>
    .app-header {
        background: linear-gradient(135deg, #1F4E79 0%, #2E75B6 100%);
        padding: 2rem 2rem 1.5rem; border-radius: 12px;
        margin-bottom: 1.5rem; text-align: center;
    }
    .app-header h1 { color: white !important; font-size: 1.9rem !important;
                     margin: 0 0 0.3rem 0 !important; }
    .app-header p  { color: #BDD7EE !important; font-size: 0.95rem !important;
                     margin: 0 !important; }
    .section-card  { background: #F8FAFD; border: 1px solid #D0DEF0;
                     border-radius: 10px; padding: 1.4rem 1.6rem;
                     margin-bottom: 1.2rem; }
    .section-title { color: #1F4E79; font-size: 1rem; font-weight: 700;
                     margin-bottom: 0.8rem; }
    .stDownloadButton > button {
        background-color: #1D6A38 !important; color: white !important;
        font-size: 1rem !important; font-weight: 600 !important;
        border-radius: 8px !important; border: none !important;
        width: 100% !important; padding: 0.65rem 1rem !important;
    }
    .stButton > button[kind="primary"] {
        background-color: #1F4E79 !important; font-size: 1.05rem !important;
        font-weight: 600 !important; border-radius: 8px !important;
        width: 100% !important; padding: 0.7rem 1.5rem !important;
    }
    [data-testid="metric-container"] {
        background: white; border: 1px solid #D0DEF0;
        border-radius: 8px; padding: 0.8rem 1rem;
    }
    .block-container { padding-top: 1.5rem !important; }
    .info-dom {
        background: #FEF3E2; border-left: 4px solid #C05621;
        border-radius: 0 8px 8px 0; padding: 0.7rem 1rem;
        font-size: 0.9rem; margin-top: 0.5rem;
        color: black !important;
    }
</style>
""", unsafe_allow_html=True)

# ── Encabezado ───────────────────────────────────────────────────────────────
st.markdown("""
<div class="app-header">
    <h1>📅 Generador de Horarios</h1>
    <p>Distribución aleatoria &nbsp;·&nbsp; Lunes a Domingo &nbsp;·&nbsp;
       Restricciones por sexo &nbsp;·&nbsp; Horario especial de domingos</p>
</div>
""", unsafe_allow_html=True)

# ── Sección 1 — Archivo ──────────────────────────────────────────────────────
st.markdown('<div class="section-card">', unsafe_allow_html=True)
st.markdown('<div class="section-title">1 · Sube tu archivo Excel</div>',
            unsafe_allow_html=True)

uploaded_file = st.file_uploader(
    "Hojas requeridas: **Actividades** y **Alumnos**",
    type=['xlsx', 'xls'],
    help="Actividades: columnas 'Nombre actividad', 'Sexo', 'Dia'.\n"
         "Alumnos: columnas 'Nombre Alumnos', 'Sexo'."
)
st.caption("Columna **Dia** en Actividades: `Toda la semana` (Lun–Dom) · `Domingo` (archivo separado)")
st.markdown('</div>', unsafe_allow_html=True)

# ── Procesar archivo ─────────────────────────────────────────────────────────
if uploaded_file is not None:
    try:
        xl = pd.ExcelFile(uploaded_file)

        for sheet in ('Actividades', 'Alumnos'):
            if sheet not in xl.sheet_names:
                st.error(f"❌ No se encontró la hoja **'{sheet}'**. "
                         f"Hojas encontradas: {', '.join(xl.sheet_names)}")
                st.stop()

        acts = xl.parse('Actividades').dropna(subset=['Nombre actividad'])
        stus = xl.parse('Alumnos').dropna(subset=['Nombre Alumnos'])

        for df, sheet, cols in [
            (acts, 'Actividades', ('Nombre actividad', 'Sexo')),
            (stus, 'Alumnos',     ('Nombre Alumnos',   'Sexo')),
        ]:
            miss = [c for c in cols if c not in df.columns]
            if miss:
                st.error(f"❌ A la hoja **'{sheet}'** le falta(n): **{', '.join(miss)}**")
                st.stop()

        # Información de las actividades por tipo de día
        has_dia_col = 'Dia' in acts.columns
        if has_dia_col:
            acts_tmp = acts.copy()
            acts_tmp['Dia'] = acts_tmp['Dia'].apply(normalize_dia)
            n_semana  = (acts_tmp['Dia'] == 'Toda la semana').sum()
            n_domingo = (acts_tmp['Dia'] == 'Domingo').sum()
        else:
            n_semana, n_domingo = len(acts), 0

        c1, c2, c3 = st.columns(3)
        with c1: st.metric("✅ Total actividades", len(acts))
        with c2: st.metric("📅 Toda la semana",    n_semana)
        with c3: st.metric("☀️ Solo domingo",      n_domingo)

        if not has_dia_col:
            st.info("ℹ️ La hoja 'Actividades' no tiene columna **Dia** — "
                    "todas las actividades se tratarán como 'Toda la semana'.")
        elif n_domingo > 0:
            st.markdown(
                '<div class="info-dom">☀️ Se generará un <b>segundo archivo Excel</b> '
                f'con las {n_domingo} actividades de domingos.</div>',
                unsafe_allow_html=True
            )

        st.divider()

        # ── Sección 2 — Tipo de horario ──────────────────────────────────────
        st.markdown('<div class="section-card">', unsafe_allow_html=True)
        st.markdown('<div class="section-title">2 · Tipo de horario</div>',
                    unsafe_allow_html=True)

        schedule_type = st.radio(
            "Periodo:", ["📅  Esta semana", "🗓️  Este mes", "📆  Todo el año"],
            horizontal=True, label_visibility="collapsed"
        )

        year  = datetime.now().year
        month = datetime.now().month

        if "mes" in schedule_type:
            col_m, col_y = st.columns([2, 1])
            with col_m:
                month_sel = st.selectbox("Mes:", MONTHS_ES,
                                         index=datetime.now().month - 1)
                month = MONTHS_ES.index(month_sel) + 1
            with col_y:
                year = int(st.number_input("Año:", min_value=2020, max_value=2060,
                                           value=datetime.now().year, step=1))
        elif "año" in schedule_type:
            year = int(st.number_input("Año:", min_value=2020, max_value=2060,
                                       value=datetime.now().year, step=1))

        st.markdown('</div>', unsafe_allow_html=True)

        # ── Sección 3 — Generar ──────────────────────────────────────────────
        st.markdown('<div class="section-card">', unsafe_allow_html=True)
        st.markdown('<div class="section-title">3 · Genera y descarga</div>',
                    unsafe_allow_html=True)

        if st.button("⚙  Generar Horario", type="primary", use_container_width=True):
            ts = datetime.now().strftime('%Y%m%d_%H%M%S')

            if "semana" in schedule_type:
                with st.spinner("Generando horario semanal…"):
                    buf_main, buf_sun, has_sun = build_weekly(acts, stus)
                name_main = f"Horario_Semanal_{ts}.xlsx"
                name_sun  = f"Domingos_Semana_{ts}.xlsx"
                resumen   = "1 semana · Lun–Dom"

            elif "mes" in schedule_type:
                mn = MONTHS_ES[month - 1]
                with st.spinner(f"Generando horario de {mn} {year}…"):
                    buf_main, buf_sun, has_sun = build_monthly(acts, stus, year, month)
                name_main = f"Horario_{mn}_{year}_{ts}.xlsx"
                name_sun  = f"Domingos_{mn}_{year}_{ts}.xlsx"
                resumen   = f"{mn} {year}"

            else:
                with st.spinner(f"Generando horario anual {year}…"):
                    buf_main, buf_sun, has_sun = build_annual(acts, stus, year)
                name_main = f"Horario_Anual_{year}_{ts}.xlsx"
                name_sun  = f"Domingos_{year}_{ts}.xlsx"
                resumen   = f"{year} · 12 hojas"

            # Guardar en session_state para que los botones persistan
            # aunque Streamlit re-ejecute el script al hacer clic en descarga
            st.session_state['horario_buf_main']  = buf_main
            st.session_state['horario_buf_sun']   = buf_sun
            st.session_state['horario_has_sun']   = has_sun
            st.session_state['horario_name_main'] = name_main
            st.session_state['horario_name_sun']  = name_sun
            st.session_state['horario_resumen']   = resumen

        # Botones de descarga — fuera del if anterior para que persistan
        # tras cualquier rerun (incluyendo el clic en el propio download_button)
        if 'horario_buf_main' in st.session_state:
            st.success(f"✅ ¡Listo! ({st.session_state['horario_resumen']})")

            has_sun = st.session_state['horario_has_sun']
            dl1, dl2 = (st.columns(2) if has_sun else (st.container(), None))

            with dl1:
                st.download_button(
                    label="📥 Horario principal (Lun–Dom)",
                    data=st.session_state['horario_buf_main'],
                    file_name=st.session_state['horario_name_main'],
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                    key="dl_main",
                )

            if has_sun and dl2:
                with dl2:
                    st.download_button(
                        label="☀️ Horario domingos",
                        data=st.session_state['horario_buf_sun'],
                        file_name=st.session_state['horario_name_sun'],
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True,
                        key="dl_sun",
                    )

        st.markdown('</div>', unsafe_allow_html=True)

    except Exception as exc:
        st.error(f"❌ Error inesperado: {exc}")
        st.exception(exc)

# ── Ayuda ────────────────────────────────────────────────────────────────────
with st.expander("ℹ️  Formato del Excel y reglas de asignación"):
    st.markdown("""
**Hoja "Actividades":**

| Nombre actividad | Sexo | Dia |
|---|---|---|
| Barrer el patio | Indiferente | Toda la semana |
| Limpiar baño H | Hombre | Toda la semana |
| Misa / Oración | Indiferente | Domingo |

**Hoja "Alumnos":** columnas `Nombre Alumnos` y `Sexo`.

**Columna Dia:**
- `Toda la semana` → aparece en el horario principal de **lunes a domingo**
- `Domingo` → aparece en un **Excel separado** que distribuye esas actividades solo el domingo

**Restricciones generales:**
- 🔵 Sexo **Hombre** → solo alumnos masculinos
- 🔴 Sexo **Mujer** → solo alumnos femeninos
- ⚪ Sexo **Indiferente** → cualquier alumno
- Cada alumno: máximo una actividad por día
- Sin repetición dentro de la misma semana (en lo posible)

**Responsable 1 / Responsable 2:**
- El sexo alterna **cada día** y **cada semana** (par = original, impar = opuesto)
- R1 y R2 siempre tienen géneros complementarios
""")

st.markdown("---")
st.caption("Generador de Horarios · Lun–Dom · Distribución aleatoria con restricciones de género")
